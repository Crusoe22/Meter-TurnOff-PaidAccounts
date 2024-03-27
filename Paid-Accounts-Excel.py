
import arcpy
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill



# Python path: C:\Program Files\ArcGIS\Pro\bin\Python\envs\arcgispro-py3 


def grabdata(): #import data from HUD_LGIM.dbo.NIGHTDUTYACCOUNTS
    # set the input feature class
    fc = r'\\GIStechnician\Share\GISDBSERVER22 GIS.sde\HUD_LGIM.dbo.NIGHTDUTYACCOUNTS'
    # set the output Excel file
    output_excel = r'\\VSERVER22\ForEveryone\CollectListReport\ShutOff.xlsx'
    # create a list of fields to extract from the feature class
    fields = ['ACCOUNTID', 'CURRTOTALDUE', 'SUMACCOUNT', 'CURRDUEDATE', 'BILLINGCYCLECODE', 'ACCOUNTNAME', 'SERVICEADDRESS', 'HOMEPHONE',  'SIZE', 'ASSETID']   
    # use the arcpy.da.SearchCursor to get all the rows in the attribute table
    rows = arcpy.da.SearchCursor(fc, fields)
    # convert the rows to a Pandas dataframe
    df = pd.DataFrame(rows, columns=fields)
    # save the dataframe to the output Excel file
    df.to_excel(output_excel, index=False)


def formatexcel(): # Format the excel sheet
    #Change header names
    # open excel file 
    dafr = pd.read_excel(r'\\VSERVER22\ForEveryone\CollectListReport\ShutOff.xlsx')

    #create copy
    correct_df = dafr.copy()

    #rename
    correct_df.rename(columns={'ACCOUNTID': 'Account ID', 'CURRTOTALDUE': 'Total Paid', 'SUMACCOUNT': 'Total Due', 'CURRDUEDATE': 'Due Date', 'BILLINGCYCLECODE': 'Billing Cycle Code', 'ACCOUNTNAME': 'Account Name', 'SERVICEADDRESS': 'Service Address', 'HOMEPHONE': 'Home Phone',
                                'SIZE': 'Size', 'ASSETID': 'Asset ID'}, inplace=True)
    

    correct_df.to_excel(r'\\VSERVER22\ForEveryone\CollectListReport\ShutOff.xlsx', index=False,header=True)
    
    color_width()


def color_width():
        # Load the workbook
        wb = openpyxl.load_workbook(r'\\VSERVER22\ForEveryone\CollectListReport\ShutOff.xlsx') # openpyxl.load_workbook(

        # Select the active worksheet
        ws = wb.active 


        # Set the width of all columns to 22
        for col in ws.columns:
            col_width = 22
            col[0].column_letter
            ws.column_dimensions[col[0].column_letter].width = col_width

        # Set the fill color for cells where SUMACCOUNT column equals 0
        for rows in ws.iter_cols(min_col=3, max_col=3, min_row=2, max_row=None):
            for cell in rows:
                if cell.value == 0:
                    cell.fill = PatternFill(start_color="00CCFFCC", end_color="00CCFFCC", fill_type = "solid") 
                elif cell.value != 0: 
                    cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")


        # Save the workbook
        wb.save(r'\\VSERVER22\ForEveryone\CollectListReport\ShutOff.xlsx')


grabdata()
formatexcel()

# print a message to confirm that the operation was successful
# print('Data exported to Excel successfully.')
